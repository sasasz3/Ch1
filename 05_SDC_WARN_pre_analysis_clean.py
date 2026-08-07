from datetime import date, datetime, time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from config import DATA_FOLDER


#

#load data
REPURCHASES_INPUT = DATA_FOLDER / "sdc_data_filtered.xlsx"


LAYOFFS_INPUT = DATA_FOLDER  / "warn_compustat_filtered.xlsx"


REPURCHASES_OUTPUT = DATA_FOLDER / "repurchases_data_final.xlsx"


LAYOFFS_OUTPUT =  DATA_FOLDER / "layoffs_data_final.xlsx"



# standardise date format
EXCEL_DATE_FORMAT = "yyyy-mm-dd"


#standardise column names
def standardise_column_name(value):
    """
    Convert a column name to lowercase and standardise the
    gvkey and date names.
    """
    if value is None:
        return None

    column_name = str(value).strip().lower()

    # Replace line breaks and repeated whitespace with one space.
    column_name = " ".join(
        column_name.split()
    )

    gvkey_names = {
        "gvkey",
        "compustat_gvkey",
    }

    if column_name in gvkey_names:
        return "gvkey"

    date_names = {
        "date announced",
        "effective date",
    }

    if column_name in date_names:
        return "date"

    return column_name

#standardise dates
def convert_to_excel_date(
    value,
    workbook_epoch,
):
    """
    Convert a cell value into a Python datetime at midnight.

    Unrecognised values return None and remain unchanged.
    """
    if value is None or value == "":
        return None

    # already a datetime.
    if isinstance(value, datetime):
        return value.replace(
            hour=0,
            minute=0,
            second=0,
            microsecond=0,
            tzinfo=None,
        )

    # already a date without a time.
    if isinstance(value, date):
        return datetime.combine(
            value,
            time.min,
        )

    # handle an Excel date serial number.
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(
                value,
                epoch=workbook_epoch,
            )

            if isinstance(converted, datetime):
                return converted.replace(
                    hour=0,
                    minute=0,
                    second=0,
                    microsecond=0,
                    tzinfo=None,
                )

            if isinstance(converted, date):
                return datetime.combine(
                    converted,
                    time.min,
                )
        except (TypeError, ValueError, OverflowError):
            return None

    # Handle date strings.
    try:
        converted = pd.to_datetime(
            value,
            errors="coerce",
        )
    except (TypeError, ValueError):
        return None

    if pd.isna(converted):
        return None

    converted = converted.to_pydatetime()

    # remove timezone information if present.
    if converted.tzinfo is not None:
        converted = converted.replace(
            tzinfo=None
        )

    return converted.replace(
        hour=0,
        minute=0,
        second=0,
        microsecond=0,
    )


def standardise_date_column(
    worksheet,
    column_number,
    workbook_epoch,
):
    """
    Convert all valid observations in one date column into
    Excel dates displayed as YYYY-MM-DD.
    """
    converted_count = 0
    blank_count = 0
    unrecognised_count = 0

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        cell = worksheet.cell(
            row=row_number,
            column=column_number,
        )

        original_value = cell.value

        if original_value is None or original_value == "":
            blank_count += 1
            continue

        # Do not replace Excel formulas.
        if (
            isinstance(original_value, str)
            and original_value.startswith("=")
        ):
            unrecognised_count += 1
            continue

        converted_value = convert_to_excel_date(
            original_value,
            workbook_epoch,
        )

        if converted_value is None:
            # Leave unrecognised values unchanged.
            unrecognised_count += 1
            continue

        cell.value = converted_value
        cell.number_format = EXCEL_DATE_FORMAT

        converted_count += 1

    return {
        "converted": converted_count,
        "blank": blank_count,
        "unrecognised": unrecognised_count,
    }


#execute in workbook

def standardise_workbook(
    input_file,
    output_file,
):
    """
    Standardise all first-row column names and every column
    whose resulting name is 'date'.
    """
    if not input_file.exists():
        raise FileNotFoundError(
            f"Input file does not exist: {input_file}"
        )

    print()
    print("=" * 75)
    print(f"Loading: {input_file.name}")
    print("=" * 75)

    workbook = load_workbook(
        input_file
    )

    total_headers_changed = 0
    total_dates_converted = 0
    total_dates_unrecognised = 0

    for worksheet in workbook.worksheets:
        print(
            f"Processing sheet: {worksheet.title}",
            flush=True,
        )

        date_column_numbers = []
        changed_columns = []

        # Standardise all headers in row 1.
        for cell in worksheet[1]:
            original_name = cell.value

            if original_name is None:
                continue

            standardised_name = (
                standardise_column_name(
                    original_name
                )
            )

            if standardised_name != original_name:
                changed_columns.append(
                    (
                        original_name,
                        standardised_name,
                    )
                )

                cell.value = standardised_name
                total_headers_changed += 1

            # Detect the standardized date column.
            if standardised_name == "date":
                date_column_numbers.append(
                    cell.column
                )

        for old_name, new_name in changed_columns:
            print(
                f"  Header: {old_name!r} -> {new_name!r}",
                flush=True,
            )

        # Standardise all identified date columns.
        for column_number in date_column_numbers:
            result = standardise_date_column(
                worksheet=worksheet,
                column_number=column_number,
                workbook_epoch=workbook.epoch,
            )

            total_dates_converted += result[
                "converted"
            ]

            total_dates_unrecognised += result[
                "unrecognised"
            ]

            print(
                f"  Date column {column_number}: "
                f"{result['converted']:,} converted, "
                f"{result['blank']:,} blank, "
                f"{result['unrecognised']:,} unrecognised.",
                flush=True,
            )

    print(
        f"Saving: {output_file.name}",
        flush=True,
    )

    workbook.save(
        output_file
    )

    print()
    print(f"Headers changed:      {total_headers_changed:,}")
    print(f"Dates standardized:   {total_dates_converted:,}")
    print(f"Dates unrecognised:   {total_dates_unrecognised:,}")
    print(f"Output file:          {output_file}")

    return {
        "headers_changed": total_headers_changed,
        "dates_converted": total_dates_converted,
        "dates_unrecognised": total_dates_unrecognised,
        "output_file": output_file,
    }


# ============================================================
# PROCESS REPURCHASES
# ============================================================

repurchases_result = standardise_workbook(
    input_file=REPURCHASES_INPUT,
    output_file=REPURCHASES_OUTPUT,
)


# ============================================================
# PROCESS LAYOFFS
# ============================================================

layoffs_result = standardise_workbook(
    input_file=LAYOFFS_INPUT,
    output_file=LAYOFFS_OUTPUT,
)


# ============================================================
# FINAL SUMMARY
# ============================================================

print()
print("=" * 75)
print("STANDARDISATION COMPLETED")
print("=" * 75)
print(
    f"Repurchases output: "
    f"{repurchases_result['output_file']}"
)
print(
    f"Layoffs output:     "
    f"{layoffs_result['output_file']}"
)
print("=" * 75)